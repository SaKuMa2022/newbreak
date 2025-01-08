import streamlit as st
import openpyxl

# Function to create Excel sheet if it doesn't exist
def create_excel_sheet():
    try:
        wb = openpyxl.load_workbook("students.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Student ID", "Name", "Age", "Sex", "Grade (Subject 1)", "Grade (Subject 2)", "Grade (Subject 3)", "Grade (Subject 4)"])
        wb.save("students.xlsx")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

# Function to submit data to Excel
def submit_data(student_id, name, age, sex, grade1, grade2, grade3, grade4):
    wb = openpyxl.load_workbook("students.xlsx")
    ws = wb.active
    ws.append([student_id, name, age, sex, grade1, grade2, grade3, grade4])
    wb.save("students.xlsx")

# Streamlit UI
st.title('Student Information Form')

# Initialize session state for input fields
if 'student_id' not in st.session_state:
    st.session_state.student_id = ""
    st.session_state.name = ""
    st.session_state.age = ""
    st.session_state.sex = "Male"
    st.session_state.grade1 = ""
    st.session_state.grade2 = ""
    st.session_state.grade3 = ""
    st.session_state.grade4 = ""

# Input fields for student data
st.session_state.student_id = st.text_input("Student ID", value=st.session_state.student_id)
st.session_state.name = st.text_input("Name", value=st.session_state.name)
st.session_state.age = st.text_input("Age", value=st.session_state.age)
st.session_state.sex = st.selectbox("Sex", ["Male", "Female", "Other"], index=["Male", "Female", "Other"].index(st.session_state.sex))
st.session_state.grade1 = st.text_input("Grade for Subject 1", value=st.session_state.grade1)
st.session_state.grade2 = st.text_input("Grade for Subject 2", value=st.session_state.grade2)
st.session_state.grade3 = st.text_input("Grade for Subject 3", value=st.session_state.grade3)
st.session_state.grade4 = st.text_input("Grade for Subject 4", value=st.session_state.grade4)

# Submit button
if st.button('Submit'):
    if (st.session_state.student_id and st.session_state.name and 
        st.session_state.age and st.session_state.sex and
        st.session_state.grade1 and st.session_state.grade2 and
        st.session_state.grade3 and st.session_state.grade4):
        
        # Save the data to Excel
        submit_data(st.session_state.student_id, st.session_state.name, st.session_state.age,
                    st.session_state.sex, st.session_state.grade1, st.session_state.grade2,
                    st.session_state.grade3, st.session_state.grade4)
        
        # Show success message
        st.success('Data submitted successfully!')

        # Clear the inputs (reset session state variables)
        st.session_state.student_id = ""
        st.session_state.name = ""
        st.session_state.age = ""
        st.session_state.sex = "Male"
        st.session_state.grade1 = ""
        st.session_state.grade2 = ""
        st.session_state.grade3 = ""
        st.session_state.grade4 = ""
    else:
        st.error('Please fill in all fields.')

# Initialize the Excel sheet on the first run
create_excel_sheet()
st.download_button(
    label="Download student data as Excel file",
    data=excel_file,
    file_name="students.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
